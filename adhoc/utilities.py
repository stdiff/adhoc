"""
Helper functions
"""

from typing import Union
from collections import OrderedDict
from itertools import product
from pathlib import Path
import tempfile
import shutil
import re

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns

from sklearn import datasets
from sklearn.utils import Bunch
from sklearn.tree import DecisionTreeClassifier, DecisionTreeRegressor

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, colors
from openpyxl.worksheet.table import Table, TableStyleInfo


class TempDir:
    """
    Helper class to work under a temporary directory.

    When an instance of this class is created, a temporary directory is
    also automatically created and you can find the path as an attribute
    temp_dir.

    Because of the implementation of tempfile.mkdtemp we have always to
    remove the temporary directory manually. You can do this by executing
    the method close().

    This class can be used as a context manager. At the end of the context
    the temporary directory is deleted.
    """

    def __init__(self):
        self._temp_dir = tempfile.mkdtemp() # type: str

    @property
    def temp_dir(self) -> Path:
        return Path(self._temp_dir)

    def __enter__(self) -> Path:
        return self.temp_dir

    def close(self):
        shutil.rmtree(self._temp_dir)
        self._temp_dir = None

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()


def bunch2dataframe(bunch:Bunch, target:str=None) -> pd.DataFrame:
    """
    Create a DataFrame from a Bunch instance. The Bunch instance must
    have attributes data (numpy array), feature_names (iterable) and
    target (numpy array or list).

    :param bunch: Bunch instance
    :param target: name of the target variable.
    :return: DataFrame
    """

    target = "target" if target is None else target
    df = pd.DataFrame(bunch.data, columns=bunch.feature_names)
    df[target] = bunch.target
    return df


def load_iris(target:str="species") -> pd.DataFrame:
    """
    Construct a DataFrame from sklearn.datasets.load_iris

    :param target: name of the target variable
    :return: DataFrame of data set
    """

    iris = datasets.load_iris()
    df = bunch2dataframe(iris, target)
    df.columns = [c[:-5].replace(" ","_") for c in iris.feature_names] + [target]
    df[target] = df[target].apply(lambda i: iris.target_names[i])
    return df


def load_breast_cancer(target:str="label") -> pd.DataFrame:
    """
    Construct a Dataframe from sklearn.datasets.load_breas_cancer

    :param target: name of the target variable
    :return: DataFrame of data set
    """
    breast_cancer = datasets.load_breast_cancer()
    df = bunch2dataframe(breast_cancer, target=target)
    df[target] = [breast_cancer.target_names[y].replace(" ","_") for y in df[target]]
    return df


def load_boston(target:str="PRICE"):
    """
    Construct a DataFrame from sklearn.datasets.load_boston

    :param target: name of the target variable
    :return: DataFrame
    """
    boston = datasets.load_boston()
    return bunch2dataframe(boston, target)


def load_diabetes(target:str="progression"):
    """
    Construct a DataFrame from sklearn.datasets.load_diabetes

    :param target: name of the target variable
    :return: DataFrame
    """

    diabetes = datasets.load_diabetes()
    return bunch2dataframe(diabetes, target)


def fetch_adult_dataset(csv_path:Path):
    """
    fetch the famous adult data set from UCI Machine Learning Repository
    and store it to the given path. If the data set file already exists,
    then the checksum is checked.

    If this function ends with no error, you have the file in the given path.

    :param csv_path: Path instance of the csv file
    """

    import hashlib
    data_url = "http://archive.ics.uci.edu/ml/machine-learning-databases/adult/adult.data"
    names = ["age", "workclass", "fnlwgt", "education", "education-num", "marital-status",
             "occupation", "relationship", "race", "sex", "capital-gain", "capital-loss",
             "hours-per-week", "native-country", "label"]
    checksum = "ee2d7503652f28a713aa6a054f5d9bb610a160afb8b817e6347d155c80af9795"

    if not csv_path.exists():
        ## if there is no data file, then we have to download the data
        df = pd.read_csv(data_url, na_values="?", names=names, skipinitialspace=True)
        df.to_csv(csv_path, index=False)

    with csv_path.open("rb") as fo:
        checksum_calculated = hashlib.sha256(fo.read()).hexdigest()

    if checksum_calculated != checksum:
        raise Exception("You seem to have downloaded a wrong file")


def grep_data(data:pd.DataFrame, column:str, expr:str) -> pd.DataFrame:
    """
    Pick the rows with a specified expression and return them as
    the subset of the given DataFrame.

    :param data: panda's DataFrame
    :param column: column to check
    :param expr: expression to find (passed to re.search)
    :return: copy of the matched rows
    """
    s_matched = data[column].apply(lambda s: True if re.search(expr,str(s)) else False)
    return data[s_matched].copy()


def facet_grid_scatter_plot(data:pd.DataFrame, row:str, col:str,
                            x:str, y:str, c:str=None, hue:str=None,
                            cmap:str="bwr", alpha=0.5, aspect:float=2,
                            margin_titles:bool=True, **kwargs):
    """
    create grid by using two categorical variables row and col and draw scatter plots by
    using variables x and y with color c (continuous) or hue (categorical). In R this
    function is something similar to

        ggplot(data, aes(x,y,color=c)) + geom_point() + facet_grid(row~col)

    (if c is continuous) or

        ggplot(data, aes(x,y,color=hue)) + geom_point() + facet_grid(row~col)

    (if c is not continuous).

    :param data: pandas DataFrame
    :param row: field of rows in a grid. The field must be categorical.
    :param col: field of columns in a grid. The field must be categorical.
    :param x: field of x-axis. This field must be continuous.
    :param y: field of y-axis. This field must be continuous.
    :param c: field of color. This field must be continuous. If hue is given, c must be None.
    :param hue: field of color. This field must be categorical. If c is given, hue must be None.
    :param cmap: parameter of matplotlib.pyplot.scatter
    :param alpha: parameter of matplotlib.pyplot.scatter
    :param aspect: parameter of seaborn.FacetGrid
    :param margin_titles: parameter of seaborn.FacetGrid
    :param kwargs: given to seaborn.FacetGrid
    :return:
    """

    def plt_scatter(x, y, c, **kwargs):
        """
        Remove color option to avoid conflict which seaborn creates.
        This idea is originated from
        https://stackoverflow.com/questions/44641669/scatterplot-with-point-colors-representing-a-continuous-variable-in-seaborn-face
        """
        kwargs.pop("color")
        plt.scatter(x, y, c=c, **kwargs)

    if c:
        ## c is given => continuous variable
        fg = sns.FacetGrid(data=data, row=row, col=col,
                           aspect=aspect, margin_titles=margin_titles, **kwargs)
        fg = fg.map(plt_scatter, x, y, c, alpha=alpha, cmap=cmap)

        ## put a common color bar on the right
        vmin, vmax = data[c].min(), data[c].max()

        fg.fig.subplots_adjust(right=0.85)
        cax = fg.fig.add_axes([0.90, 0.25, 0.02, 0.6]) ## TODO: is it universal?
        points = plt.scatter([], [], c=[], vmin=vmin, vmax=vmax, cmap=cmap)
        color_bar = fg.fig.colorbar(points, cax=cax)
        color_bar.set_label(c, rotation=270, labelpad=25)

    else:
        ## hue is given => discrete variable
        fg = sns.FacetGrid(data=data, row=row, col=col, hue=hue,
                           aspect=aspect, margin_titles=margin_titles, **kwargs)
        fg.map(plt.scatter, x, y, alpha=alpha).add_legend()


def bins_by_tree(data:pd.DataFrame, field:str, target:str,
                 target_is_continuous:bool, n_bins:int=5,
                 n_points:int=200, precision:int=2, **kwargs) -> pd.Series:
    """
    bin the given field by looking at the target variable. More precisely
    we bin the field, so that the cost function (Gini index/RMSE) is optimized.
    For implementation we just train a decision tree model.

    WARNING: the number of bins can be smaller than n_bins, because of the training result.

    :param data: DataFrame containing field and target
    :param field: column to bin. Must be continuous.
    :param target: column which is used to calculate cost function.
    :param target_is_continuous: True if target variable is continuous.
    :param n_bins: number of bins.
    :param n_points: number of points of grid to check
    :param precision: parameter of pandas.cut
    :param kwargs: parameter for DecisionTreeRegressor/Classifier
    :return: binned Series
    """

    if target_is_continuous:
        tree = DecisionTreeRegressor(max_leaf_nodes=n_bins, **kwargs)
    else:
        tree = DecisionTreeClassifier(max_leaf_nodes=n_bins, **kwargs)

    ## TODO: cross-validation (?)
    tree.fit(data[[field]], data[target])

    ## grid space whose points can be separators
    grid = np.linspace(data[field].min(), data[field].max(), num=n_points)
    grid = pd.DataFrame({field:grid})

    if target_is_continuous:
        fibers = pd.DataFrame(tree.predict(grid))
    else:
        fibers = pd.DataFrame(tree.predict_proba(grid))

    ## possible values of predictions
    leaves = fibers.drop_duplicates().copy()
    leaves["bin"] = np.arange(0,leaves.shape[0]) ## put bin numbers

    grid = pd.merge(pd.concat([grid,fibers], axis=1), leaves, how="left", on=list(fibers.columns))

    ## find "bins" (boundaries of bins)
    bins = list(grid.groupby("bin")[field].first())
    bins[0] = -np.inf
    bins.append(np.inf)

    return pd.cut(data[field], bins, precision=precision)


def bins_heatmap(data:pd.DataFrame, cat1:str, cat2:str, x:str, y:str, target:str,
                 n_bins:int=5, cmap:str="YlGnBu", center:float=None,
                 magnification:float=None, fontsize:int=14):
    """
    create a grid with cat1 (and cat2) and draw heat maps (x,y,AVG(target)). For heat maps
    we bin the variables x and y by applying bins_by_tree.

    - cat1 and cat2 must be categorical
    - x and y must be continuous and they are binned.
    - target variable must be numerical, so that the average has meaning.

    :param data: DataFrame
    :param cat1: categorical variable for a grid.
    :param cat2: Another categorical variable. You can put None, if you do not need it.
    :param x: continuous variable
    :param y: continuous variable
    :param target: numerical variable.
    :param n_bins: *maximum* number of bins. The actual number can be smaller than it.
    :param cmap: Parameter of seaborn.heatmap.
    :param center: center of the color gauge.
    :param magnification: magnification of the height.
    :param fontsize: font size of title, xlabel and ylabel
    """

    cats1 = data[cat1].unique()

    if cat2 is None:
        cats2 = [None]
    else:
        cats2 = data[cat2].unique()

    vals2loc = OrderedDict()
    for val1, val2 in product(cats1,cats2):
        if val2 is None:
            loc = data[cat1] == val1
        else:
            loc = np.logical_and(data[cat1] == val1, data[cat2] == val2)

        if loc.any():
            vals2loc[(val1,val2)] = loc

    fig, axes = plt.subplots(nrows=len(vals2loc.keys()), ncols=1)

    ## change the height of the plot
    h = fig.get_figheight()
    if magnification is None:
        magnification = max(len(cats1), len(cats2))
    fig.set_figheight(magnification*h)

    ## boundary of the color gauge
    vmin, vmax = data[target].min(), data[target].max()

    for pair, subax in zip(vals2loc.keys(), axes.flat):
        val1, val2 = pair
        subset = data.loc[vals2loc[(val1,val2)],:].copy()

        for field in [x,y]:
            subset[field] = bins_by_tree(subset, field=field, target=target,
                                         target_is_continuous=True, n_bins=n_bins)

        df_heatmap = subset.groupby([x,y])[target].mean().reset_index()
        df_heatmap = df_heatmap.pivot(index=y, columns=x, values=target)
        df_heatmap.sort_index(ascending=False, inplace=True)
        sns.heatmap(df_heatmap, vmin=vmin, vmax=vmax, cmap=cmap,
                    center=center, annot=True, cbar=False, ax=subax)

        if val2 is None:
            subax.set_title("%s = %s" % (cat1, val1))
        else:
            subax.set_title("%s = %s | %s = %s" % (cat1, val1, cat2, val2))

        subax.title.set_fontsize(fontsize)
        subax.set_xlabel(x, fontsize=fontsize)
        subax.set_ylabel(y, fontsize=fontsize)

    plt.tight_layout()


def to_excel(df:pd.DataFrame, file:Union[str,Path], sheet:str="Sheet",
             libreoffice:bool=True, style:str="TableStyleMedium9"):
    """
    Save the given DataFrame into an Excel file.

    The reason why we do not use pd.DataFrame.to_excel is the style.
    Basically we give a style to the table through openpyxl, but
    The given style can be ignored on LibreOffice. Therefore we set
    the standard colors to the cells if libreoffice option is True.
    But in this case, we ignore the option style.

    The index will be ignored.
    TODO: option whether the index should also be written

    :param df: DataFrame
    :param file: path to the excel file (xlsx)
    :param sheet: name of the sheet for the table
    :param libreoffice: arrange style for libreOffice
    :param style: name of the style. If libreoffice is True, this is ignored.
    """

    if not Path(file).exists():
        ## if the file hat not been created we choose the default sheet
        wb = Workbook()
        ws = wb.active
        ws.title = sheet

    else:
        wb = load_workbook(str(file))
        if sheet in wb.sheetnames:
            ## if the specified sheet exits, then we overwrite it.
            ws_index = wb.sheetnames.index(sheet)
            wb.remove(wb[sheet])
            ws = wb.create_sheet(title=sheet, index=ws_index)

        else:
            ## otherwise we create a new sheet.
            ws = wb.create_sheet(title=sheet)

    ## header
    ws.append(df.columns.tolist())

    ## values
    for i,row in enumerate(df.iterrows()):
        ws.append(row[1].tolist())

    ## Direct settings for LibreOffice
    ## We can choose a style of a table through TableStyleInfo,
    ## but it is not applied in LibreOffice.
    if libreoffice:
        style = "TableStyleMedium9"

        color_header = colors.Color(rgb='4f81bd')
        colors_row = ['b8cce4', 'dce6f1']
        for i in range(1,df.shape[0]+2):
            color = color_header if i == 1 else colors_row[i % 2]

            for j in range(1,df.shape[1]+1):
                if i == 1:
                    ## header (bold, white)
                    ws.cell(column=j, row=i).font = Font(bold=True, color="FFFFFF")

                ws.cell(column=j, row=i).fill = PatternFill(fgColor=color, fill_type="solid")

    ## table
    ## https://openpyxl.readthedocs.io/en/stable/worksheet_tables.html
    last_cell = "%s%s" % (ws.cell(row=df.shape[0]+1, column=df.shape[1]).column_letter,
                          df.shape[0]+1)
    tab = Table(displayName="Table1", ref="A1:%s" % last_cell)
    style = TableStyleInfo(name=style, showFirstColumn=False, showLastColumn=False,
                           showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    wb.save(filename=str(file))

    
if __name__ == "__main__":
    pass
