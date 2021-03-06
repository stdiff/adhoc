"""
unittest for adhoc/utilities.py
"""

from unittest import TestCase

import numpy as np
import pandas as pd
from sklearn import datasets

from adhoc.utilities import TempDir
from adhoc.utilities import bunch2dataframe, fetch_adult_dataset
from adhoc.utilities import load_iris, load_boston, load_breast_cancer, load_diabetes
from adhoc.utilities import grep_data, bins_by_tree, to_excel
from adhoc.utilities import facet_grid_scatter_plot, bins_heatmap

import warnings
warnings.filterwarnings("ignore")

class TestUtilities(TestCase):

    def setUp(self):
        np.random.seed(1)


    def test_buynch2dataframe(self):
        """unittest for bunch2dataframe"""
        data = datasets.load_iris()
        target = "label"
        df = bunch2dataframe(data, target=target)

        self.assertTrue(isinstance(df,pd.DataFrame))
        self.assertEqual(list(df.columns)[-1], target)


    def test_load_iris(self):
        """unittest for load_iris"""
        target = "species"
        cats = {"setosa","versicolor","virginica"}
        df = load_iris(target=target)

        self.assertTrue(isinstance(df,pd.DataFrame))
        self.assertEqual((150,5), df.shape)
        self.assertEqual(target, list(df.columns)[-1])
        self.assertEqual(cats,set(df[target].unique()))


    def test_load_boston(self):
        """unittest for load_iris"""
        target = "target"
        df = load_boston(target=target)

        self.assertTrue(isinstance(df,pd.DataFrame))
        self.assertEqual((506,14), df.shape)
        self.assertEqual(target, list(df.columns)[-1])


    def test_load_breast_cancer(self):
        """unittest for load_breast_cancer"""
        target = "label"
        df = load_breast_cancer(target=target)

        self.assertTrue(isinstance(df,pd.DataFrame))
        self.assertEqual((569,31), df.shape)
        self.assertEqual(target, list(df.columns)[-1])


    def test_load_diabetes(self):
        """unittest for load_boston"""
        target = "label"
        df = load_diabetes(target=target)

        self.assertTrue(isinstance(df,pd.DataFrame))
        self.assertEqual((442,11), df.shape)
        self.assertEqual(target, list(df.columns)[-1])


    def test_fetch_adult_dataset(self):
        with TempDir() as tmp_dir:
            dummy_file = tmp_dir.joinpath("dummy.csv")
            with dummy_file.open("w") as fo:
                fo.write("1,2,3\n")

            ## wrong file -> Exception
            with self.assertRaises(Exception):
                fetch_adult_dataset(dummy_file)


    def test_grep_data(self):
        df = pd.DataFrame({
            "col1": list("abcdefg"),
            "col2": [2,3,5,7,11,13,17]
        })

        df_selected = grep_data(df,"col1","[a-c]")
        self.assertIsInstance(df_selected,pd.DataFrame)
        self.assertEqual(3, df_selected.shape[0])

        ## specified column will be converted in str
        df_selected = grep_data(df,"col2",r"^\d$")
        self.assertIsInstance(df_selected,pd.DataFrame)
        self.assertEqual(4, df_selected.shape[0])


    def test_bins_by_tree(self):
        """unittest for bins_by_tree"""
        df = load_iris(target="species")
        cols = list(df.columns)
        n_bins = 3

        ## Case) The target variable is continuous
        bins = bins_by_tree(df, field=cols[2], target=cols[3],
                            target_is_continuous=True,
                            n_bins=n_bins, n_points=200, precision=1)
        cats = sorted(bins.unique())

        self.assertIsInstance(bins, pd.Series)
        self.assertEqual(len(bins.unique()), n_bins)

        for cat in cats:
            self.assertIsInstance(cat, pd.Interval)
            self.assertEqual(cat.closed,"right")

        self.assertEqual(cats[0].left, -np.inf)
        self.assertEqual(cats[-1].right, np.inf)

        ## Case) The target variable is not continuous
        bins = bins_by_tree(df, field=cols[2], target=cols[-1],
                            target_is_continuous=False,
                            n_bins=n_bins, n_points=200, precision=1)
        cats = sorted(bins.unique())

        self.assertIsInstance(bins, pd.Series)
        self.assertEqual(len(bins.unique()), n_bins)

        for cat in cats:
            self.assertIsInstance(cat, pd.Interval)
            self.assertEqual(cat.closed,"right")

        self.assertEqual(cats[0].left, -np.inf)
        self.assertEqual(cats[-1].right, np.inf)


    def test_facet_grid_scatter_plot(self):
        ## Check if it works without any error.
        np.random.seed(1)

        df = load_iris(target="species")
        df["grid_col"] = np.random.choice(["c1","c2"],
                                          size=df.shape[0],
                                          replace=True)
        df["grid_row"] = np.random.choice(["r1","r2","r3"],
                                          size=df.shape[0],
                                          replace=True)

        facet_grid_scatter_plot(data=df,
                                row="grid_row", col="grid_col",
                                x="petal_width", y="petal_length",
                                c="sepal_width")

        facet_grid_scatter_plot(data=df,
                                row="grid_row", col="grid_col",
                                x="petal_width", y="petal_length",
                                hue="species")


    def test_bins_heatmap(self):
        # check if this function works without error
        np.random.seed(2)
        df = load_iris(target="species")
        df["cat1"] = np.random.choice(["a","b"],
                                      size=df.shape[0],
                                      replace=True)

        bins_heatmap(df, cat1="cat1", cat2="species",
                     x="petal_width", y="petal_length",
                     target="sepal_length",
                     n_bins=3)


    def test_to_excel(self):
        df_iris = load_iris()
        df_boston = load_boston()
        df_cancer = load_breast_cancer().loc[:10,:]

        from openpyxl import load_workbook
        sheets = ["iris","boston"]

        with TempDir() as temp_dir:
            excel_file = temp_dir.joinpath("new_file.xlsx")

            ## Case 1) Create a new file
            to_excel(df=df_iris, file=excel_file, sheet=sheets[0], libreoffice=True)

            wb1 = load_workbook(str(excel_file))
            self.assertEqual([sheets[0]], wb1.sheetnames)

            ## whether we can read the excel through pandas.read_excel
            dg1 = pd.read_excel(excel_file, sheet_name=sheets[0])
            self.assertEqual(df_iris.columns.tolist(),dg1.columns.tolist())
            self.assertEqual(df_iris.shape, dg1.shape)

            ## Case 2) Add a sheet to an existing excel file
            to_excel(df=df_boston, file=excel_file, sheet=sheets[1], libreoffice=True)

            wb2 = load_workbook(str(excel_file))
            self.assertEqual(sheets, wb2.sheetnames)

            dg2 = pd.read_excel(excel_file, sheet_name=sheets[1])
            self.assertEqual(df_boston.columns.tolist(),dg2.columns.tolist())
            self.assertEqual(df_boston.shape, dg2.shape)

            ## Case 3) Overwrite an existing sheet with a new table
            ## This table has no style.
            to_excel(df=df_cancer, file=excel_file, sheet=sheets[0], libreoffice=False)

            wb3 = load_workbook(str(excel_file))
            self.assertEqual(sheets, wb3.sheetnames)

            dg3 = pd.read_excel(excel_file, sheet_name=sheets[0])
            self.assertEqual(df_cancer.columns.tolist(), dg3.columns.tolist())
            self.assertEqual(df_cancer.shape, dg3.shape)

            # If you want to see the result
            #from shutil import copy
            #copy(str(excel_file), "/tmp/test_to_excel.xlsx")
